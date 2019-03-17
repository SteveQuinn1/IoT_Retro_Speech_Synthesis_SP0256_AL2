# Copyright 2017 Steve Quinn
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
#  This program is distributed in the hope that it will be useful,
#  but WITHOUT ANY WARRANTY; without even the implied warranty of
#  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#  GNU General Public License for more details.
#
#  You should have received a copy of the GNU General Public License
#  along with this program.  If not, see <http://www.gnu.org/licenses/>.
#
#  Filename : Test1.py
#  Purpose : To automatically test an IoT device via MQTT publications and send results to a log file
#
#  Uses python v3.5.2
#  Found here : https://www.python.org/downloads/release/python-352/
#
#  Requires the following modules;
#  paho-mqtt 1.3.0, Found here : https://pypi.python.org/pypi/paho-mqtt/1.3.0
#  OpenPyXL 2.5.0a3, Found here : https://pypi.python.org/pypi/openpyxl
#
#  Usage
#  -----
#  Command line
#  >python Test1.py filename [v]
#
#  python : Is the path to your python v3.5.2 instalation
#  Test1.py : This file
#  filename : Is the input filename in the form below.
#  [v] : Verbose option. If the letter v is added, data sent to log file will also be sent to the the system console
#
#  Input filename
#  --------------
#  Can be single spreadsheet.xlsx, or a filename.txt, where the text file contains multiple spreadsheets.xlsx
#  The format of the spread sheet is:
#  There must be the following tabs in the spread sheet
#      Config, Preamble, Postamble, Testx
#      Test tabs must start with the name 'Test'
#

import os
from openpyxl import load_workbook
import sys
import time
import datetime
import paho.mqtt.client as mqtt
import socket

# fCompound
compound_file_flag = bool(False)
compound_file = str("")
compound_file_plus_path = str("")
compound_file_terminator = str(".txt")
compound_file_content_terminator = str(".xlsx")
compound_file_source_file = str("")
compound_file_source_file_plus_path = str("")

# Source spread sheet key tab names
preconditions_tabname = str("Preconditions")
config_tabname = str("Config")
preamble_tabname = str("Preamble")
postamble_tabname = str("Postamble")
test_tabname_prefix = str("Test")

STR_PASS = "Pass"
STR_FAIL = "Fail"

total_tests = 0
total_test_passes = 0
total_test_fails = 0

dir_path = os.path.dirname(os.path.realpath(__file__))

# Config
MQTT_IP_Address = str("")
MQTT_IP_Port = str("")
Log_Path = str("")
Default_Log_Path = dir_path
Log_Filename = str("")
log_file_terminator = str(".txt")
Default_Log_Filename = str("Logfile.txt")
System_timer = float(0.0)
Default_System_timer = float(60000.0)



# fLog
log_file = Default_Log_Filename
log_file_plus_path = dir_path + '\\' + log_file
fLogFilehandle = None
bSendToSystemConsole = False
tStart = None
tCurrent = None
tNew = None
tOld = None
# These are shared between 'def handle_config(ws, sLogfilename)' and 'def handle_spreadsheet(spreadsheet_filename_and_path)'
VALUE_COLUMN = int(2)
LOG_PATH_ROW = int(4)
LOG_FILENAME_ROW = int(5)

# fSource
source_file = str("")
source_file_plus_path = str("")





# The callback for when the client receives a CONNACK response from the broker.
def on_connect(client, userdata, flags, rc):
    global bMQTTConnected
    bMQTTConnected = True
    strTemp = "Connected to MQTT Broker, RC " + str(rc)
    WriteToLog("> " + strTemp)



def on_disconnect(client, userdata, rc):
#def on_disconnect(client, userdata, flags, rc):
    global bMQTTConnected
    bMQTTConnected = False

    strTemp = "Disconnected from MQTT Broker, RC " + str(rc)
    WriteToLog("> " + strTemp)



# The callback for when a PUBLISH message is received from the broker.
def on_message(client, userdata, msg):
    #print(msg.topic+" "+str(msg.payload))
    global ExpectedSubscriptionTopic
    global bMQTTRecievedExpectedPublication
    global ReceivedSubscriptionTopicPayload
    sTopic = str(msg.topic)
    #sPayload = str(msg.payload)
    sPayload = str(msg.payload.decode("utf-8"))
    WriteToLog("  Received Topic/Payload " + sTopic + " '" + sPayload + "'\n")
    if ExpectedSubscriptionTopic.find(sTopic) >= 0:
        bMQTTRecievedExpectedPublication = True
        ReceivedSubscriptionTopicPayload = sPayload


# MQTT
bFirstTime = True
client = mqtt.Client()
client.on_connect = on_connect
client.on_disconnect = on_disconnect
client.on_message = on_message
bMQTTConnected = False
bMQTTRecievedExpectedPublication = False
ReceivedSubscriptionTopicPayload = str("")
ExpectedSubscriptionTopic = str("")

MSG_INSTRUCTION_COL = int(1)
MSG_INSTRUCTION_ROW_OFFSET = int(2)
MSG_PUBLISH_TOPIC_COL = int(2)
MSG_PUBLISH_PAYLOAD_COL = int(3)
MSG_SUBSCRIPTION_TOPIC_COL = int(4)
MSG_SUBSCRIPTION_PAYLOAD_COL = int(5)
MSG_MESSAGE_DELAY_COL = int(6)
MSG_MESSAGE_WHEN_COL = int(7)
MSG_MESSAGE_COMMENT_COL = int(8)
MSG_DELAY_VALUE_COL = int(2)
MSG_REMARK_COL = int(2)


MWT_BEFORE_TIME_T = "B"
MWT_AFTER_TIME_T = "A"
MWT_NOTHING_FOR_TIME_T = "NFT"
MWT_NOTHING_BEFORE_TIME_T = "NBT"
MWT_SOMETHING_BEFORE_TIME_T = "SBT"

message_when_types = [MWT_BEFORE_TIME_T, MWT_AFTER_TIME_T, MWT_NOTHING_FOR_TIME_T, MWT_NOTHING_BEFORE_TIME_T, MWT_SOMETHING_BEFORE_TIME_T]


def message_handler( ws, current_row ):
    global bMQTTRecievedExpectedPublication
    global ReceivedSubscriptionTopicPayload
    global ExpectedSubscriptionTopic

    PublishTopic = ""
    PublishPayload = ""
    SubscriptionTopic = ""
    SubscriptionPayload = ""
    MessageDelay = ""
    MessageWhen = ""
    MessageComment = ""

    if (ws.cell(row=current_row, column=MSG_MESSAGE_WHEN_COL).value != None):
        MessageWhen = str(ws.cell(row=current_row, column=MSG_MESSAGE_WHEN_COL).value)
        if ((ws.cell(row=current_row, column=MSG_PUBLISH_TOPIC_COL).value != None) or ((ws.cell(row=current_row, column=MSG_PUBLISH_TOPIC_COL).value == None) and ((MessageWhen.upper().find(MWT_SOMETHING_BEFORE_TIME_T) >= 0) or (MessageWhen.upper().find(MWT_NOTHING_BEFORE_TIME_T) >= 0)))):
            if (ws.cell(row=current_row, column=MSG_PUBLISH_TOPIC_COL).value != None):
                PublishTopic = str(ws.cell(row=current_row, column=MSG_PUBLISH_TOPIC_COL).value)
            if ((ws.cell(row=current_row, column=MSG_PUBLISH_PAYLOAD_COL).value != None) or ((ws.cell(row=current_row, column=MSG_PUBLISH_PAYLOAD_COL).value == None) and ((MessageWhen.upper().find(MWT_SOMETHING_BEFORE_TIME_T) >= 0) or (MessageWhen.upper().find(MWT_NOTHING_BEFORE_TIME_T) >= 0)))):
                if (ws.cell(row=current_row, column=MSG_PUBLISH_PAYLOAD_COL).value != None):
                    PublishPayload = str(ws.cell(row=current_row, column=MSG_PUBLISH_PAYLOAD_COL).value)
                    if PublishPayload.find("*") >= 0:
                        PublishPayload = ""
                if (ws.cell(row=current_row, column=MSG_SUBSCRIPTION_TOPIC_COL).value != None):
                    SubscriptionTopic = str(ws.cell(row=current_row, column=MSG_SUBSCRIPTION_TOPIC_COL).value)
                    if (ws.cell(row=current_row, column=MSG_SUBSCRIPTION_PAYLOAD_COL).value != None):
                        SubscriptionPayload = str(ws.cell(row=current_row, column=MSG_SUBSCRIPTION_PAYLOAD_COL).value)
                        if (ws.cell(row=current_row, column=MSG_MESSAGE_DELAY_COL).value != None):
                            MessageDelay = str(ws.cell(row=current_row, column=MSG_MESSAGE_DELAY_COL).value)
                            if MessageWhen in message_when_types:
                                if (ws.cell(row=current_row, column=MSG_MESSAGE_COMMENT_COL).value != None):
                                    MessageComment = str(ws.cell(row=current_row, column=MSG_MESSAGE_COMMENT_COL).value)
                                bMQTTRecievedExpectedPublication = False
                                ReceivedSubscriptionTopicPayload = ""
                                ExpectedSubscriptionTopic = SubscriptionTopic
                                bSystemTimerExpiry = False
                                bMessageDelayTimerExpiry = False
                                bTerminateLoop = False

                                tNewSystemTimer = time.time()
                                tOldSystemTimer = tNewSystemTimer
                                tNewMessageDelay = tNewSystemTimer
                                tOldMessageDelay = tNewMessageDelay
                                tElapsedMessageDelay = float(0.0)
                                sPassFail = "X"

                                if (MessageWhen.upper().find(MWT_SOMETHING_BEFORE_TIME_T) >= 0) or (MessageWhen.upper().find(MWT_NOTHING_BEFORE_TIME_T) >= 0):
                                    strTemp = "Published Message/Payload : " + MessageWhen.upper() + " Has no trigger event [Line " + str(current_row) + "]"
                                else:
                                    strTemp = "Published Message/Payload : " + PublishTopic + " " + PublishPayload + " [Line " + str(current_row) + "]"
                                WriteToLog("  " + strTemp)
                                if (MessageWhen.upper().find(MWT_SOMETHING_BEFORE_TIME_T) < 0) and (MessageWhen.upper().find(MWT_NOTHING_BEFORE_TIME_T) < 0):
                                    client.publish(PublishTopic, PublishPayload, 0, False)
                                while not (bSystemTimerExpiry or bTerminateLoop):
                                    tNewSystemTimer = time.time()
                                    if (tNewSystemTimer - tOldSystemTimer) >= System_timer:
                                        bSystemTimerExpiry = True
                                    tNewMessageDelay = tNewSystemTimer
                                    tElapsedMessageDelay = tNewMessageDelay - tOldMessageDelay
                                    bMessageDelayTimerExpiry = ((float(MessageDelay)/1000.0) < tElapsedMessageDelay)

                                    if MessageWhen.upper().find(MWT_BEFORE_TIME_T) >= 0:
                                        if (not bTerminateLoop) and bMQTTRecievedExpectedPublication and (not bMessageDelayTimerExpiry):
                                            sPassFail = "P"
                                            bTerminateLoop = True

                                        if (not bTerminateLoop) and bMessageDelayTimerExpiry and (not bMQTTRecievedExpectedPublication):
                                            sPassFail = "F"
                                            bTerminateLoop = True

                                    elif MessageWhen.upper().find(MWT_AFTER_TIME_T) >= 0:
                                        if (not bTerminateLoop) and bMQTTRecievedExpectedPublication and bMessageDelayTimerExpiry:
                                            sPassFail = "P"
                                            bTerminateLoop = True

                                        if (not bTerminateLoop) and bMQTTRecievedExpectedPublication and (not bMessageDelayTimerExpiry):
                                            sPassFail = "F"
                                            bTerminateLoop = True
                                    # Duplicate case
                                    elif MessageWhen.upper().find(MWT_NOTHING_FOR_TIME_T) >= 0:
                                        if (not bTerminateLoop) and bMessageDelayTimerExpiry  and (not bMQTTRecievedExpectedPublication):
                                            sPassFail = "P"
                                            bTerminateLoop = True

                                        if (not bTerminateLoop) and (not bMessageDelayTimerExpiry)  and bMQTTRecievedExpectedPublication:
                                            sPassFail = "F"
                                            bTerminateLoop = True

                                    elif MessageWhen.upper().find(MWT_NOTHING_BEFORE_TIME_T) >= 0:
                                        if (not bTerminateLoop) and bMessageDelayTimerExpiry and (not bMQTTRecievedExpectedPublication):
                                            sPassFail = "P"
                                            bTerminateLoop = True

                                        if (not bTerminateLoop) and (not bMessageDelayTimerExpiry) and bMQTTRecievedExpectedPublication:
                                            sPassFail = "F"
                                            bTerminateLoop = True

                                    elif MessageWhen.upper().find(MWT_SOMETHING_BEFORE_TIME_T) >= 0:
                                        if (not bTerminateLoop) and (not bMessageDelayTimerExpiry) and bMQTTRecievedExpectedPublication:
                                            sPassFail = "P"
                                            bTerminateLoop = True

                                        if (not bTerminateLoop) and bMessageDelayTimerExpiry and (not bMQTTRecievedExpectedPublication):
                                            sPassFail = "F"
                                            bTerminateLoop = True

                                if bSystemTimerExpiry:
                                    strTemp = "System timer expired " + str(current_row)
                                    WriteToLog("  " + strTemp)
                                    return STR_FAIL

                                if MessageWhen.upper().find(MWT_BEFORE_TIME_T) >= 0:
                                    strTemp = "Expecting '" + SubscriptionPayload + "'"
                                    WriteToLog("  " + strTemp)
                                    if sPassFail == "F":
                                        return STR_FAIL
                                    elif sPassFail == "P":
                                        if SubscriptionPayload.lower() == ReceivedSubscriptionTopicPayload.lower():
                                            return STR_PASS
                                        elif SubscriptionPayload == "*":
                                            return STR_PASS
                                        else:
                                            return STR_FAIL
                                    else:
                                        return STR_FAIL

                                elif MessageWhen.upper().find(MWT_AFTER_TIME_T) >= 0:
                                    strTemp = "Expecting " + SubscriptionPayload
                                    WriteToLog("  " + strTemp)
                                    if sPassFail == "F":
                                        return STR_FAIL
                                    elif sPassFail == "P":
                                        if SubscriptionPayload.lower() == ReceivedSubscriptionTopicPayload.lower():
                                            return STR_PASS
                                        elif SubscriptionPayload == "*":
                                            return STR_PASS
                                        else:
                                            return STR_FAIL
                                    else:
                                        return STR_FAIL

                                elif MessageWhen.upper().find(MWT_NOTHING_FOR_TIME_T) >= 0:
                                    if sPassFail == "F":
                                        return STR_FAIL
                                    elif sPassFail == "P":
                                        return STR_PASS

                                if MessageWhen.upper().find(MWT_NOTHING_BEFORE_TIME_T) >= 0:
                                    if sPassFail == "F":
                                        return STR_FAIL
                                    elif sPassFail == "P":
                                        return STR_PASS

                                elif MessageWhen.upper().find(MWT_SOMETHING_BEFORE_TIME_T) >= 0:
                                    strTemp = "Expecting " + SubscriptionPayload
                                    WriteToLog("  " + strTemp)
                                    if sPassFail == "F":
                                        return STR_FAIL
                                    elif sPassFail == "P":
                                        if SubscriptionPayload.lower() == ReceivedSubscriptionTopicPayload.lower():
                                            return STR_PASS
                                        elif SubscriptionPayload == "*":
                                            return STR_PASS
                                        else:
                                            return STR_FAIL
                                    else:
                                        return STR_FAIL

                            else:
                                strTemp = "Unrecognised Message When " + str(current_row)
                                WriteToLog("  " + strTemp)
                        else:
                            strTemp = "Missing Message Delay " + str(current_row)
                            WriteToLog("  " + strTemp)
                    else:
                        strTemp = "Missing Subscription Payload " + str(current_row)
                        WriteToLog("  " + strTemp)
                else:
                    strTemp = "Missing Subscription Topic " + str(current_row)
                    WriteToLog("  " + strTemp)
            else:
                strTemp = "Missing Publish Payload " + str(current_row)
                WriteToLog("  " + strTemp)
        else:
            strTemp = "Missing Publish Topic " + str(current_row)
            WriteToLog("  " + strTemp)
    else:
        strTemp = "Missing Message Sign " + str(current_row)
        WriteToLog("  " + strTemp)  #strTemp = "Out message_handler " + str(current_row)
    #WriteToLog("  " + strTemp)
    return STR_FAIL



def delay_handler( ws, current_row ):
    DelayValue = ""
    if (ws.cell(row=current_row, column=MSG_DELAY_VALUE_COL).value != None):
        DelayValue = str(ws.cell(row=current_row, column=MSG_DELAY_VALUE_COL).value)
        strTemp = "Delay : " + DelayValue + "mS [Line " + str(current_row) + "]"
        WriteToLog("  " + strTemp)
        time.sleep(float(DelayValue)/1000.0)
    else:
        strTemp = "Missing Delay Value " + str(current_row)
        WriteToLog("  " + strTemp)

    #strTemp = "In  delay_handler " + str(row)
    #WriteToLog("> " + strTemp)
    return STR_PASS


def bcast_message_handler( ws, current_row ):
    PublishTopic = ""
    PublishPayload = ""
    if (ws.cell(row=current_row, column=MSG_PUBLISH_TOPIC_COL).value != None):
        PublishTopic = str(ws.cell(row=current_row, column=MSG_PUBLISH_TOPIC_COL).value)

        if (ws.cell(row=current_row, column=MSG_PUBLISH_PAYLOAD_COL).value != None):
            PublishPayload = str(ws.cell(row=current_row, column=MSG_PUBLISH_PAYLOAD_COL).value)
            client.publish(PublishTopic, PublishPayload, 0, False)
            strTemp = "Published Message/Payload : " + PublishTopic + " " + PublishPayload + " [Line " + str(current_row) + "]"
            WriteToLog("  " + strTemp)
            return STR_PASS
        else:
            strTemp = "Missing Publish Payload " + str(current_row)
            WriteToLog("  " + strTemp)
    else:
        strTemp = "Missing Publish Topic " + str(current_row)
        WriteToLog("  " + strTemp)

    #strTemp = "In  delay_handler " + str(row)
    #WriteToLog("> " + strTemp)
    return STR_FAIL


def remark_handler( ws, current_row ):
    RemarkText = ""
    if (ws.cell(row=current_row, column=MSG_REMARK_COL).value != None):
        RemarkText = str(ws.cell(row=current_row, column=MSG_REMARK_COL).value)
        strTemp = "Remark : " + RemarkText + " [Line " + str(current_row) + "]"
        WriteToLog("  " + strTemp)
    else:
        strTemp = "Missing Remark"
        WriteToLog("  " + strTemp)
    return STR_PASS



def external_handler( ws, row ):
    strTemp = "In  external_handler " + str(row)
    WriteToLog("> " + strTemp)
    return False



REMARK_TOKEN = "Rem"
MESSAGE_TOKEN = "Message"
BCAST_MESSAGE_TOKEN = "BCastMessage"
DELAY_TOKEN = "Delay"
EXTERNAL_TOKEN = "External"
instruction_handler_array = {REMARK_TOKEN: remark_handler, BCAST_MESSAGE_TOKEN: bcast_message_handler, MESSAGE_TOKEN: message_handler, DELAY_TOKEN : delay_handler, EXTERNAL_TOKEN : external_handler}
subscription_array = list()




def handle_test(ws):
    #global total_tests
    #global total_test_passes
    #global total_test_fails

    bTestPassed = True

    WriteToLog("> " + ws.title + " Start\n")
    if ws.title.lower().find(test_tabname_prefix.lower()) == 0:
        row_count = ws.max_row
        for row_position in range(MSG_INSTRUCTION_ROW_OFFSET, row_count + 1):
            if (ws.cell(row=row_position, column=MSG_INSTRUCTION_COL).value != None):
                strTemp = str(ws.cell(row=row_position, column=MSG_INSTRUCTION_COL).value)
                bInstructionFound = False
                for instruction, instruction_handler in instruction_handler_array.items():
                    if strTemp.lower().find(instruction.lower()) == 0:
                        TestResult = instruction_handler_array[instruction]( ws, row_position )
                        bInstructionFound = True
                        if bTestPassed and (TestResult.find(STR_FAIL)>=0):
                            bTestPassed = False
                if not bInstructionFound:
                    strTemp = "Unrecognised instruction [" + strTemp + "] in test " + ws.title + " at line " + str(row_position) + "\n"
                    WriteToLog("  " + strTemp)
    if bTestPassed:
        strTemp = ws.title + " End [" + STR_PASS + "]\n\n"
        WriteToLog("> " + strTemp)
        return STR_PASS
    else:
        strTemp = ws.title + " End [" + STR_FAIL + "]\n\n"
        WriteToLog("> " + strTemp)
        return STR_FAIL


def handle_subsriptions(wb):
    for ws in wb.worksheets:
        if ws.title.lower().find(test_tabname_prefix.lower()) == 0:
            row_count = ws.max_row
            for index in range(1, row_count+1):
                #strTemp = str(ws.cell(row=index, column=MSG_INSTRUCTION_COL).value)
                if (ws.cell(row=index, column=MSG_INSTRUCTION_COL).value != None) and (str(ws.cell(row=index, column=MSG_INSTRUCTION_COL).value).lower().find(MESSAGE_TOKEN.lower()) >= 0):
                    if (ws.cell(row=index, column=MSG_SUBSCRIPTION_TOPIC_COL).value != None):
                        if (subscription_array == None) or not (str(ws.cell(row=index, column=MSG_SUBSCRIPTION_TOPIC_COL).value) in subscription_array):
                            subscription_array.append(str(ws.cell(row=index, column=MSG_SUBSCRIPTION_TOPIC_COL).value))
    if bMQTTConnected:
        for topic in subscription_array:
            client.subscribe(topic)
            time.sleep(0.5)
            if subscription_array.index(topic) != len(subscription_array) - 1:
                strTemp = "Subscribing to [" + topic + "]"
            else:
                strTemp = "Subscribing to [" + topic + "]\n\n"
            WriteToLog("> " + strTemp)
    else:
        strTemp = "No broker connection. Unable to subscribe"
        WriteToLog("> " + strTemp)



def count_tests(wb):
    test_count = 0
    for ws in wb.worksheets:
        if ws.title.lower().find(test_tabname_prefix.lower()) == 0:
            test_count += 1
    return test_count



def handle_config(ws, sLogfilename):
    errNo = 0
    global MQTT_IP_Address
    global MQTT_IP_Port
    global Log_Path
    global Log_Filename
    global System_timer
    global log_file_plus_path
    global VALUE_COLUMN
    MQTT_IP_ADDRESS_ROW = 2
    MQTT_PORT_ROW = 3
    global LOG_PATH_ROW
    global LOG_FILENAME_ROW
    SYSTEM_TIMER_ROW = 6
    WriteToLog("> Config Start\n")
    # Read MQTT_IP_Address
    MQTT_IP_Address = ws.cell(row=MQTT_IP_ADDRESS_ROW, column=VALUE_COLUMN).value
    if MQTT_IP_Address != None:
        MQTT_IP_Address = MQTT_IP_Address.strip(" ")
    if not (is_valid_local_address(MQTT_IP_Address) or is_valid_ipv4_address(MQTT_IP_Address) or is_valid_ipv6_address(MQTT_IP_Address) or (not (MQTT_IP_Address == None))):
        WriteToLog("  Failed to read valid MQTT_IP_Address\n")
        errNo = 1
    else :
        WriteToLog("  Using MQTT_IP_Address \'" + MQTT_IP_Address + "\'\n")

    # Read MQTT_IP_Port
    MQTT_IP_Port = int(ws.cell(row=MQTT_PORT_ROW, column=VALUE_COLUMN).value)
    if (MQTT_IP_Port == 0) :
        WriteToLog("  Failed to read valid MQTT_IP_Port\n")
        errNo = 2
    else :
        WriteToLog("  Using MQTT_IP_Port \'%s\'\n" % MQTT_IP_Port)

    # Read Log_Path
    Log_Path = ws.cell(row=LOG_PATH_ROW, column=VALUE_COLUMN).value
    if Log_Path != None:
        Log_Path = Log_Path.strip(" ")
    if (not os.path.exists(Log_Path)) or (Log_Path == None) or  (len(Log_Path) == 0) or (Log_Path == "."):
        Log_Path = Default_Log_Path
        WriteToLog("  Using default Log_Path \'" + Log_Path + "\'\n")
    else :
        WriteToLog("  Using Log_Path \'" + Log_Path + "\'\n")

    # Read Log_Filename
    Log_Filename = ws.cell(row=LOG_FILENAME_ROW, column=VALUE_COLUMN).value
    if Log_Filename != None:
        Log_Filename = Log_Filename.strip(" ")
    if (Log_Filename == None) or (len(Log_Filename) == 0) or (Log_Filename == "."):
        Log_Filename = Default_Log_Filename
        WriteToLog("  Using default Log_Filename \'" + Log_Filename + "\'\n")
    elif (Log_Filename == "*"):
        Log_Filename = sLogfilename
        WriteToLog("  Using Spreadsheet name as Log_Filename \'" + Log_Filename + "\'\n")
    else :
        WriteToLog("  Using Log_Filename \'" + Log_Filename + "\'\n")

    log_file_plus_path = Log_Path + '\\' + Log_Filename

    # Read System_timer
    if (ws.cell(row=SYSTEM_TIMER_ROW, column=VALUE_COLUMN).value == None):
        System_timer = float(Default_System_timer)/1000.0
        #print("  Using Default System Timer \'" + str(System_timer) + "\'\n")
        WriteToLog("  Using Default System Timer Value \'" + str(System_timer) + "\'\n")
    else:
        System_timer = float(ws.cell(row=SYSTEM_TIMER_ROW, column=VALUE_COLUMN).value)/1000.0
        # print("  System Timer \'" + str(System_timer) + "\'\n")
        WriteToLog("  System Timer Value \'" + str(System_timer) + "\'\n")
    WriteToLog("> Config End\n")
    return errNo



def is_valid_local_address(address):
    if ((".local" in address)) and (len(address) > len(".local")):
        return True



def is_valid_ipv4_address(address):
    try:
        socket.inet_pton(socket.AF_INET, address)
    except AttributeError:  # no inet_pton here, sorry
        try:
            socket.inet_aton(address)
        except socket.error:
            return False
        return address.count('.') == 3
    except socket.error:  # not a valid address
        return False

    return True



def is_valid_ipv6_address(address):
    try:
        socket.inet_pton(socket.AF_INET6, address)
    except socket.error:  # not a valid address
        return False
    return True



def handle_postamble(ws):
    # Not implemented as yet
    #WriteToLog("> " + ws.title + " Start\n")
    WriteToLog("> Postamble Start\n")
    #print ("In handle_postamble", end=" ")
    #print (ws.cell(row=1, column=1).value)
    #time.sleep(3)
    WriteToLog("> Postamble End\n")
    #WriteToLog("> " + ws.title + " End\n")



def handle_preamble(ws):
    # Not implemented as yet
    #WriteToLog("> " + ws.title + " Start\n")
    WriteToLog("> Preamble Start\n")
    #print ("In handle_preamble", end=" ")
    #print (ws.cell(row=1, column=1).value)
    #time.sleep(1)
    WriteToLog("> Preamble End\n")
    #WriteToLog("> " + ws.title + " End\n")



def get_worksheet_handle( wb, worksheet_name ):
    ws_id = 0
    for sheet in wb.worksheets:
        #print ("In get_worksheet_handle : ",sheet.title, ws_id)
        if sheet.title.lower() == worksheet_name.lower():
            return ws_id
        ws_id += 1
    return -1



def pend_mqtt_connection():
    global System_timer
    global bMQTTConnected
    tNewLocal = time.time()
    tOldLocal = tNewLocal

    #print(" In pend_mqtt_connection")
    while not bMQTTConnected:
        tNewLocal = time.time()
        if (tNewLocal - tOldLocal) >= System_timer:
            print(" Not Connected")
            return False
        #time.sleep(0.1)
    #print(" Connected")
    return True



def handle_spreadsheet(spreadsheet_filename_and_path):
    global Log_Path
    global Log_Filename
    global VALUE_COLUMN
    global LOG_PATH_ROW
    global LOG_FILENAME_ROW
    global total_tests
    global total_test_passes
    global total_test_fails
    global compound_file_flag
    global bFirstTime

    sPath, sFilename = os.path.split(spreadsheet_filename_and_path)
    wb = load_workbook(filename= spreadsheet_filename_and_path)
    ws_id = get_worksheet_handle( wb, config_tabname )
    if ws_id >= 0:
        # Need to read log path and filename before calling OpenLogging() to set log_file_plus_path

        # Read Log_Path
        Log_Path = wb.worksheets[ws_id].cell(row=LOG_PATH_ROW, column=VALUE_COLUMN).value
        if (Log_Path != None):
            Log_Path = Log_Path.strip(" ")
        if (not os.path.exists(Log_Path)) or (Log_Path == None) or (len(Log_Path) == 0) or (Log_Path == "."):
            Log_Path = Default_Log_Path

        # Read Log_Filename
        Log_Filename = wb.worksheets[ws_id].cell(row=LOG_FILENAME_ROW, column=VALUE_COLUMN).value
        if (Log_Filename != None):
            Log_Filename = Log_Filename.strip(" ")
        if (Log_Filename == None) or (len(Log_Filename) == 0) or (Log_Filename == "."):
            Log_Filename = Default_Log_Filename
        elif (Log_Filename == "*"):
            filename, file_extension = os.path.splitext(sFilename)
            Log_Filename = filename + log_file_terminator

        log_file_plus_path = Log_Path + '\\' + Log_Filename

        OpenLogging(False, log_file_plus_path, " " + sFilename)

        handle_config(wb.worksheets[ws_id], Log_Filename)

        ws_id = get_worksheet_handle( wb, preamble_tabname )
        if ws_id >= 0 :
            handle_preamble(wb.worksheets[ws_id])
        ws_id = get_worksheet_handle( wb, postamble_tabname )
        if  ws_id >= 0 :
            handle_postamble(wb.worksheets[ws_id])

        #client.loop_start()
        #client.connect(MQTT_IP_Address, int(MQTT_IP_Port))
        try:
            if bFirstTime :
                bFirstTime = False
                client.connect(MQTT_IP_Address, int(MQTT_IP_Port))

            if pend_mqtt_connection():
                handle_subsriptions(wb)
                total_tests = count_tests(wb)
                for sheet in wb.worksheets:
                    if sheet.title.lower().find(test_tabname_prefix.lower()) == 0:
                        strSuccessState = handle_test(sheet)
                        if strSuccessState.find(STR_PASS) >= 0:
                            total_test_passes += 1
                        elif strSuccessState.find(STR_FAIL) >= 0:
                            total_test_fails += 1
                    elif sheet.title.lower().find(test_tabname_prefix.lower()) > 0:
                        strTemp = "Test [" + sheet.title + "] ignored\n\n"
                        WriteToLog("> " + strTemp)
                #client.disconnect()
            else :
                strTemp = "Failed to make MQTT broker connection"
                WriteToLog("> " + strTemp)
            #client.loop_stop(force=False)
            #client.loop_stop(force=True)
        except:
            strTemp = "client.connect() failure. Bad IP address? " + str(MQTT_IP_Address)
            WriteToLog("> " + strTemp)
        #client.loop_stop(force=True)
        CloseLogging("")
    else :
        #print("Unable to find tab : \'" + config_tabname + "\' in : " + sFilename)
        strTemp = "Unable to find tab : \'" + config_tabname + "\' in : " + sFilename
        WriteToLog("> " + strTemp)



def EOF(f):
    current_pos = f.tell()
    file_size = os.fstat(f.fileno()).st_size
    return not (current_pos < file_size)



def OpenLogging( bPreserve, sLogfilenameAndPath , sStartingComment):
    global tStart
    global tOld
    global tNew
    global fLogFilehandle

    tStart = time.strftime("%d/%m/%Y %H:%M:%S")
    tOld = tNew = time.time()

    iIndex = 0
    if not os.path.isfile(sLogfilenameAndPath):
        fLogFilehandle = open(sLogfilenameAndPath, 'w')
    else:
        if not bPreserve :
            fLogFilehandle = open(sLogfilenameAndPath, 'w')
        else :
            while not os.path.isfile(sLogfilenameAndPath + str(iIndex)):
                iIndex += 1
            fLogFilehandle = open(sLogfilenameAndPath + str(iIndex), 'w')

    if not fLogFilehandle.closed:
        strTemp = "Logging started " + tStart + sStartingComment
        if not ("\n" in strTemp):
            strTemp += "\n"
        if bSendToSystemConsole:
            print(strTemp, end="")
        fLogFilehandle.write(strTemp)

    return fLogFilehandle



def CloseLogging( sClosingComment ):
    global tCurrent
    global fLogFilehandle
    global total_tests
    global total_test_passes
    global total_test_fails

    tCurrent = time.strftime("%H:%M:%S")
    if not fLogFilehandle.closed:
        strTemp = "Logging Ends " + tCurrent + sClosingComment + " Total Tests " + str(total_tests) + ", Passes " + str(total_test_passes) + ", Fails " + str(total_test_fails)
        if not ("\n" in strTemp):
            strTemp += "\n\n"
        fLogFilehandle.write(strTemp)
        if bSendToSystemConsole:
            print(strTemp, end="")
        fLogFilehandle.close()



def WriteToLog( sLine ):
    global fLogFilehandle
    strTemp = GetLogTimeString("", "") + sLine
    if not ("\n" in strTemp):
        strTemp += "\n"
    SendToSystemConsole(strTemp, False)
    if not fLogFilehandle == None :
        if not fLogFilehandle.closed:
            fLogFilehandle.write(strTemp)



def SendToSystemConsole( sLine, bWithTimeStamp ):
    global bSendToSystemConsole
    if bWithTimeStamp:
        strTemp = GetLogTimeString("", "") + sLine
    else:
        strTemp = sLine

    if not ("\n" in strTemp):
        strTemp += "\n"
    if bSendToSystemConsole:
        print(strTemp, end="")



def GetLogTimeString( sBefore, sAfter ):
    global tNew
    global tOld
    global tCurrent
    tCurrent = time.strftime("%H:%M:%S")
    tNew = time.time()
    tDelta = tNew - tOld
    #str = sBefore + "Start " + tStart + " Current " + tCurrent + " Delta " + FormatTime(tDelta) + sAfter
    str = sBefore + "S " + tStart + " F " + tCurrent + " D " + FormatTime(tDelta) + " " + sAfter
    tOld = tNew
    return str



def FormatTime(tTime):
    hours = tTime // 3600
    tTime = tTime - 3600 * hours
    minutes = tTime // 60
    seconds = tTime - 60 * minutes
    return "%02d:%02d:%02f" % (hours, minutes, seconds)



if len(sys.argv) >= 2:
    source_file = str(sys.argv[1])
    if len(sys.argv) == 3:
        bSendToSystemConsole = ("v" in str(sys.argv[2]).lower())
    source_file_plus_path = dir_path + '\\' + source_file
    if os.path.isfile(source_file_plus_path):
        compound_file_flag = (source_file.find(compound_file_terminator) >= 0)
        bFirstTime = True
        if not compound_file_flag :
            client.loop_start()
            handle_spreadsheet(source_file_plus_path)
        else :
            fCompound = open(source_file_plus_path,"r")
            line_count = 0
            client.loop_start()
            while not EOF(fCompound):
                compound_file_source_file = str(fCompound.readline())
                compound_file_source_file = compound_file_source_file.strip("\n")
                line_count += 1
                if len(compound_file_source_file) > 0:
                    if compound_file_source_file.find(compound_file_content_terminator) >= 0:
                        compound_file_source_file_plus_path = dir_path + '\\' + compound_file_source_file
                        if os.path.isfile(compound_file_source_file_plus_path):
                            total_tests = 0
                            total_test_passes = 0
                            total_test_fails = 0
                            handle_spreadsheet(compound_file_source_file_plus_path)
                        else :
                            print("File \'" + compound_file_source_file + "\' does not exist")
                    else :
                        print("File \'" + compound_file_source_file + "\' has wrong file type")
                else :
                    print("File \'" + source_file + "\' has empty line, at line : " + str(line_count))
            fCompound.close()
        if bMQTTConnected:
            client.disconnect()
    else :
        print("File \'" + source_file + "\' does not exist")
else :
    print("No source file name given")